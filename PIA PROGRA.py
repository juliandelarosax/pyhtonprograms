import sys
import random
from datetime import datetime, timedelta
import sqlite3 
from sqlite3 import Error 
import os
import openpyxl


def inicializar_contador():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS ContadorFolios (folio INTEGER PRIMARY KEY)")

            cursor.execute("SELECT folio FROM ContadorFolios")
            valor_actual = cursor.fetchone()

            if valor_actual is None:
                cursor.execute("INSERT INTO ContadorFolios (folio) VALUES (1)")
                conn.commit()
    except Error as e:
        print("Error al inicializar el contador de folios:", e)

def primera_ejecucion():
    return not os.path.exists("CoworkingLTI.db")

if primera_ejecucion():
    print("Bienvenido a la primera ejecución del programa")
    inicializar_contador()      

def Crear_tabla ():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn: mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Usuarios (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Salas (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, capacidad INTEGER NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Reservaciones (folio INTEGER PRIMARY KEY, nombre TEXT NOT NULL, horario Text NOT NULL, fecha timestamp) ")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS ContadorFolios (folio INTEGER PRIMARY KEY) ")
        print("Tablas creadas exitosamente")
    except Error as e:
         print (e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close() 
    
Crear_tabla ()

def menu_principal():
    while True:
        print("*******MENU PRINCIPAL*******")
        print("1. Reservaciones")
        print("2. Reportes")
        print("3. Registrar una sala")
        print("4. Registrar un nuevo cliente")
        print("5. Eliminar sala")
        print("6. Eliminar cliente")
        print("7. Salir")
        opcion = input("Ingrese el número de la opción a realizar: ")

        if opcion == "1":
            gestionar_reservaciones()
        elif opcion == "2":
            reportes()
        elif opcion == "3":
            registrar_sala()
        elif opcion == "4":
            nuevo_cliente()
        elif opcion == "5":
            eliminar_sala()
        elif opcion == "6":
            eliminar_cliente()
        elif opcion == "7":
            sys.exit()
        else:
            print("Opción no válida, inténtelo nuevamente")

def gestionar_reservaciones():
    while True:
        print("*******GESTIÓN DE RESERVACIONES*******")
        print("1. Registrar una nueva reservación")
        print("2. Editar una reserva existente")
        print("3. Consultar disponibilidad de salas")
        print("4. Eliminar una reservación")
        print("5. Volver al menú principal")

        opcion = input("Ingrese el número de la opción a realizar: ")

        if opcion == "1":
            registrar_reservacion()
        elif opcion == "2":
            editar_reservacion()
        elif opcion == "3":
            consultar_disponibilidad()
        elif opcion == "4":
            eliminar_reservacion()
        elif opcion == "5":
            return
        else:
            print("Opción no válida, inténtelo nuevamente")

def ver_disponibilidad_sala(sala, fecha, turno):
    with sqlite3.connect("CoworkingLTI.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT clave FROM Salas WHERE nombre = ? AND fecha = ? AND turno = ?", (sala, fecha, turno))
        sala_existente = cursor.fetchone()
        return sala_existente is None

def registrar_reservacion():
    print("*******REGISTRO DE RESERVACIÓN*******")
    nombre_cliente = input("Nombre del cliente: ")

    with sqlite3.connect("CoworkingLTI.db") as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT clave FROM Usuarios WHERE nombre = ?", (nombre_cliente,))
        resultado_cliente = cursor.fetchone()

        if resultado_cliente is None:
            print("El cliente no está registrado. Regístrese primero para realizar una reservación.")
            return

    nombre_evento = input("Nombre del evento: ")

    nombre_sala = input("Nombre de la sala: ")

    capacidad_sala = int(input("Capacidad de la sala: "))

    fecha = input("Fecha de la reserva (AAAA-MM-DD): ")

    turno = input("Escriba el turno deseado (mañana, tarde, noche): ")

    if not nombre_evento or not nombre_sala or not fecha or not turno:
        print("Debe llenar todos los campos, son obligatorios")
        return

    if not ver_disponibilidad_sala(nombre_sala, fecha, turno):
        print("La sala no está disponible en la fecha y horario deseados")
        return

    cursor.execute("SELECT * FROM Reservaciones WHERE nombre = ? AND fecha = ? AND horario = ? AND nombre = ?",
                   (nombre_evento, fecha, turno, nombre_sala))
    reservacion_existente = cursor.fetchone()

    if reservacion_existente:
        print("La sala ya está reservada para el mismo horario y fecha. Elija otra sala, fecha u horario.")
        return

    fecha2 = datetime.now().strftime("%Y-%m-%d")
    fecha_reservacion = datetime.strptime(fecha, "%Y-%m-%d")
    fecha_actual = datetime.strptime(fecha2, "%Y-%m-%d")
    anticipacion = fecha_reservacion - fecha_actual

    if anticipacion.days < 2:
        print("La reservación debe hacerse con al menos 2 días de anticipación")
        return

    try:
        conn.execute("BEGIN")

        cursor.execute("SELECT MAX(folio) FROM Reservaciones")
        max_folio = cursor.fetchone()[0]
        clave_reserva = max_folio + 1 if max_folio is not None else 1

        clave_aleatoria = random.randint(1, 99999)

        cursor.execute("INSERT INTO Salas (clave, nombre, capacidad, fecha, turno) VALUES (?, ?, ?, ?, ?)",
                       (clave_aleatoria, nombre_sala, capacidad_sala, fecha, turno))

        cursor.execute("INSERT INTO Reservaciones (folio, nombre, horario, fecha) VALUES (?, ?, ?, ?)",
                       (clave_reserva, nombre_evento, turno, fecha))

        conn.commit()

        print("La reservación ha sido registrada con éxito. Su número de folio es:", clave_reserva)
        print("El número clave para su sala es:", clave_aleatoria)
    except sqlite3.Error as e:
        conn.rollback()
        print("Error al registrar la reservación:", e)

def generar_folio_unico():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT folio FROM ContadorFolios")
            resultado = cursor.fetchone()

            if resultado:
                valor_actual = resultado[0]
                nuevo_valor = valor_actual + 1
                cursor.execute("UPDATE ContadorFolios SET folio = ?", (nuevo_valor,))
                conn.commit()
                return nuevo_valor
            else:
                nuevo_valor = 1
                cursor.execute("INSERT INTO ContadorFolios (folio) VALUES (?)", (nuevo_valor,))
                conn.commit()
                return nuevo_valor
    except Error as e:
        print("Error al generar el folio único:", e)
        return None
    
def editar_reservacion():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()

            folio_a_editar = input("Ingrese el folio de la reservación que desee editar: ")
            clave_sala = input("Ingrese la clave de la sala: ")

            cursor.execute("SELECT * FROM Reservaciones WHERE folio = ?", (folio_a_editar,))
            reserva = cursor.fetchone()

            cursor.execute("SELECT * FROM Salas WHERE clave = ?", (clave_sala,))
            sala = cursor.fetchone()

            if not reserva or not sala:
                print("Debe llenar todos los campos, son obligatorios")
                return

            if not reserva:
                print("La reservación con el folio ingresado no existe o es incorrecta. Intente nuevamente.")
                return

            if not sala:
                print("La clave de sala no es válida o no coincide con la sala de la reserva. Intente nuevamente.")
                return
            
            if not reserva or not sala:
                print("Debe llenar todos los campos, son obligatorios")
                return

            print("Detalles de la reservación a editar:")
            print("")
            print("Folio:", reserva[0])
            print("")
            print("Nombre del evento:", reserva[1])
            print("")
            print("Horario:", reserva[2])
            print("")
            print("Fecha:", reserva[3])
            print("")
            print("Detalles de la sala:")
            print("")
            print("Nombre de la sala:", sala[1])
            print("")
            print("Capacidad de la sala:", sala[2])
            print("")
            print("Fecha de la sala:", sala[3])
            print("")
            print("Turno de la sala:", sala[4])

            nuevo_folio = input("Nuevo folio de la reservación (deje en blanco para mantener el actual): ")
            nuevo_clave_sala = input("Nueva clave de la sala (deje en blanco para mantener la actual): ")
            nuevo_nombre_evento = input("Nuevo nombre del evento (deje en blanco para mantener el actual): ")
            nuevo_nombre_sala = input("Nuevo nombre de la sala (deje en blanco para mantener el actual): ")
            nuevo_horario = input("Nuevo horario (deje en blanco para mantener el actual): ")
            nueva_fecha = input("Nueva fecha (AAAA-MM-DD) (deje en blanco para mantener la actual): ")

            if (
                nuevo_folio
                or nuevo_clave_sala
                or nuevo_nombre_evento
                or nuevo_nombre_sala
                or nuevo_horario
                or nueva_fecha
            ):
                query_reservacion = "UPDATE Reservaciones SET "
                query_sala = "UPDATE Salas SET "
                values_reservacion = []
                values_sala = []

                if nuevo_folio:
                    query_reservacion += "folio = ?, "
                    values_reservacion.append(nuevo_folio)
                if nuevo_nombre_evento:
                    query_reservacion += "nombre = ?, "
                    values_reservacion.append(nuevo_nombre_evento)
                if nuevo_horario:
                    query_reservacion += "horario = ?, "
                    values_reservacion.append(nuevo_horario)
                if nueva_fecha:
                    query_reservacion += "fecha = ?, "
                    values_reservacion.append(nueva_fecha)

                if nuevo_clave_sala:
                    query_sala += "clave = ?, "
                    values_sala.append(nuevo_clave_sala)
                if nuevo_nombre_sala:
                    query_sala += "nombre = ?, "
                    values_sala.append(nuevo_nombre_sala)

                query_reservacion = query_reservacion.rstrip(', ') + " WHERE folio = ?"
                values_reservacion.append(folio_a_editar)

                query_sala = query_sala.rstrip(', ') + " WHERE clave = ?"
                values_sala.append(clave_sala)

                cursor.execute(query_reservacion, tuple(values_reservacion))
                cursor.execute(query_sala, tuple(values_sala))
                conn.commit()

                print("Reservación y sala actualizadas con éxito")
            else:
                print("No se realizaron cambios en la reserva y la sala")
    except Error as e:
        print("Error al editar la reserva:", e)



def consultar_disponibilidad():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor= conn.cursor()

            fecha_consulta= input("Ingrese la fecha (AAAA-MM-DD) para consultar disponibilidad: ")

            cursor.execute("SELECT clave, nombre, turno FROM Salas "
                           "WHERE clave NOT IN (SELECT DISTINCT clave FROM Reservaciones "
                           "WHERE fecha = ?) AND fecha = ?",
                           (fecha_consulta,fecha_consulta))
            
            salas_disponibles= cursor.fetchall()

            if not salas_disponibles:
                print("No hay salas disponibles para esa fecha")
            else:
                print("Salas disponibles en la fecha", fecha_consulta)
                for sala in  salas_disponibles:
                    print(f"Clave: {sala[0]}, Nombre: {sala[1]}, Turno: {sala[2]}")
    except Error as e:
        print("Error al consultar disponibilidad:", e)
    return

def actualizar_contador_folios():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM Reservaciones")
            cantidad_reservaciones = cursor.fetchone()[0]

            cursor.execute("UPDATE ContadorFolios SET folio = ?", (cantidad_reservaciones,))
            conn.commit()
    except Error as e:
        print("Error al actualizar el contador de folios:", e)
    
def eliminar_reservacion():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()

            folio_a_eliminar = input("Ingrese el folio de la reserva a eliminar: ")
            clave_sala = input("Ingrese la clave de la sala: ")

            cursor.execute("SELECT * FROM Reservaciones WHERE folio = ?", (folio_a_eliminar,))
            reserva = cursor.fetchone()

            cursor.execute("SELECT clave FROM Salas WHERE clave = ?", (clave_sala,))
            sala = cursor.fetchone()

            if not folio_a_eliminar or not clave_sala:
                print("Debe llenar todos los campos, son obligatorios")
                return

            if not reserva:
                print("La reservación con el folio ingresado no existe.")
                return

            if not sala:
                print("La clave de sala no es válida o no coincide con la sala de la reserva.")
                return
                    

            print("Detalles de la reservación a eliminar:")
            print("")
            print("Folio: ", reserva[0])
            print("")
            print("Nombre del evento: ", reserva[1])
            print("")
            print("Horario:", reserva[2])
            print("")
            print("Fecha:", reserva[3])
            print("")
            print("Detalles de la sala a eliminar:")
            print("")
            print("Clave de la sala:", reserva[1])
            print("")
            print("Nombre de la sala:", reserva[2])
            print("")
            print("Turno de la sala:", reserva[3])

            confirmacion = input("¿Está seguro de que desea eliminar la reservación? (S/N): ")

            if confirmacion.upper() == "S":
                cursor.execute("DELETE FROM Reservaciones WHERE folio = ?", (folio_a_eliminar,))
                conn.commit()
                actualizar_contador_folios()

                cursor.execute("DELETE FROM Salas WHERE clave = ?", (clave_sala,))
                conn.commit()

                print("La reservación ha sido eliminada exitosamente, y la sala también ha sido eliminada.")
            else:
                print("La reservación no ha sido eliminada.")
    except Error as e:
        print("Error al eliminar la reservación:", e)

def reportes():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()

            fecha_reporte = input("Ingrese la fecha (AAAA-MM-DD) para generar el reporte: ")

            ##fecha_reporte_timestamp = datetime.strptime(fecha_reporte, "%Y-%m-%d")

            cursor.execute("SELECT Reservaciones.nombre, Salas.nombre, Usuarios.nombre, Reservaciones.horario, Reservaciones.fecha FROM Reservaciones, Salas, Usuarios WHERE date(Reservaciones.fecha) = ? AND Reservaciones.nombre = Salas.nombre AND Reservaciones.horario = Salas.turno AND Reservaciones.nombre = Usuarios.nombre",
                            (fecha_reporte,))
            reservaciones = cursor.fetchall()

            if not reservaciones:
                print("No hay reservaciones para la fecha proporcionada.")
                return

            print(f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_reporte}")
            print("{:<20} {:<20} {:<20} {:<20} {:<20}".format("Sala", "Cliente", "Evento", "Turno", "Horario"))
            for reserva in reservaciones:
                print("{:<20} {:<20} {:<20} {:<20} {:<20}".format(reserva[0], reserva[1], reserva[2], reserva[3], reserva[4]))

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"Reporte Reservaciones {fecha_reporte}"

            headers = ["Sala", "Cliente", "Evento", "Turno", "Horario"]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header

            for row_num, reserva in enumerate(reservaciones, 2):
                for col_num, valor in enumerate(reserva, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = valor

            archivo_excel = f"reporte_reservaciones_{fecha_reporte}.xlsx"
            workbook.save(archivo_excel)

            print(f"El reporte ha sido exportado a '{archivo_excel}'.")
    except Error as e:
        print("Error al generar el reporte:", e)

def registrar_sala():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor = conn.cursor()

            nombre_cliente = input("Nombre del cliente que registra la sala: ")
            cursor.execute("SELECT clave FROM Usuarios WHERE nombre = ? ", (nombre_cliente,))

            resultado = cursor.fetchone()

            if not nombre_cliente:
                print("Debe llenar todos los campos, son obligatorios")
                return

            if resultado is None:
                print("El cliente no está registrado. Debe registrarse primero para registrar una sala.")
                return

            cursor.execute("SELECT MAX(clave) FROM Salas")
            max_clave = cursor.fetchone()[0]
            clave_sala = max_clave + 1 if max_clave is not None else 1

            nombre_sala = input("Ingrese el nombre de la sala: ")

            fecha_sala = input("Ingrese la fecha que usará la sala (AAAA-MM-DD): ") 

            turno = input("Ingrese el turno deseado (mañana, tarde, noche): ")

            cupo_sala = input("Ingrese el cupo de la sala: ")

            if clave_sala < 1:
                print("La clave debe ser un número positivo")
                return
            if not nombre_sala or not turno or not fecha_sala: 
                print("Debe llenar todos los campos, son obligatorios")
                return

            cursor.execute("INSERT INTO Salas (clave, nombre, fecha, turno, capacidad) VALUES (?, ?, ?, ?, ?)",
                           (clave_sala, nombre_sala, fecha_sala, turno, cupo_sala))
            conn.commit()

            print(f"La sala con clave {clave_sala} ha sido registrada exitosamente.")

    except sqlite3.Error as e:
        print("Error al registrar la sala: ", e)

def nuevo_cliente():
    try:
        with sqlite3.connect("CoworkingLTI.db") as conn:
            cursor= conn.cursor()

            clave_cliente= input("Ingrese la clave unica del cliente: ")

            nombre_cliente= input("Ingrese el  nombre del cliente: ")
            
            
            if not clave_cliente or not nombre_cliente:
                print("Debe llenar todos los campos, son obligatorios")
                return
            
            if clave_cliente < "1":
                print("La clave debe ser un numero consecuente de 0")
                return
            
            cursor.execute("INSERT INTO Usuarios (clave, nombre) VALUES (?, ?)",
                           (clave_cliente, nombre_cliente))
            conn.commit()

            print("El cliente ha sido registrado exitosamente")

    except Error as e:

        print("Error al registrar al cliente: ", e)

def eliminar_sala():
    print("*******ELIMINAR SALA*******")
    nombre_cliente = input("Nombre del cliente: ")

    with sqlite3.connect("CoworkingLTI.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT clave FROM Usuarios WHERE nombre = ?", (nombre_cliente,))

        resultado = cursor.fetchone()

        if resultado is None:
            print("El cliente no está registrado. Regístrese primero para obtener el permiso.")
            return

    clave_sala = int(input("Clave de la sala a eliminar: "))

    cursor.execute("SELECT clave, nombre, horario FROM Salas WHERE clave = ?", (clave_sala,))

    sala = cursor.fetchone()

    if sala is None:
        print("La sala con la clave ingresada no existe.")
        return

    print("Detalles de la sala a eliminar:")

    print("Clave de la sala:", sala[0])

    print("Nombre de la sala:", sala[1])

    print("Horario de la sala:", sala[2])

    confirmacion = input("¿Está seguro de que desea eliminar la sala? (S/N): ")

    if confirmacion.upper() == "S":
        cursor.execute("DELETE FROM Salas WHERE clave = ?", (clave_sala,))
        conn.commit()
        print("La sala se ha borrado correctamente.")
    else:
        print("La sala no ha sido eliminada.")
        return   

def eliminar_cliente():
    print("*******ELIMINAR CLIENTE*******")
    nombre_cliente = input("Nombre del cliente registrado: ")

    with sqlite3.connect("CoworkingLTI.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT clave FROM Usuarios WHERE nombre = ?", (nombre_cliente,))

        resultado = cursor.fetchone()

        if resultado is None:
            print("El cliente no está registrado. No puede acceder a esta parte del programa.")
            return

    nombre_usuario = input("Nombre del cliente a eliminar: ")

    cursor.execute("SELECT clave, nombre FROM Usuarios WHERE nombre = ?", (nombre_usuario,))

    usuario = cursor.fetchone()

    if usuario is None:
        print("El cliente con el nombre ingresado no existe.")
        return

    print("Detalles del cliente a eliminar:")

    print("Clave del cliente:", usuario[0])

    print("Nombre del cliente:", usuario[1])

    confirmacion = input("¿Está seguro de que desea eliminar al cliente? (S/N): ")

    if confirmacion.upper() == "S":
        cursor.execute("DELETE FROM Reservaciones WHERE nombre = ?", (usuario[1],))
        cursor.execute("DELETE FROM Salas WHERE horario = ?", (usuario[0],))
        cursor.execute("DELETE FROM ContadorFolios WHERE folio = ?", (usuario[0],))
        cursor.execute("DELETE FROM Usuarios WHERE clave = ?", (usuario[0],))
        conn.commit()
        print("El cliente ha sido eliminado exitosamente, junto con sus reservaciones, salas y folios relacionados.")
    else:
        print("El cliente no ha sido eliminado.")


menu_principal()
            




      